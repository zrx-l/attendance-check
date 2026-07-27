import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)

import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
import re
import argparse
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.chart import BarChart, PieChart, Reference

# ================== 颜色常量 ==================
GREEN_RGB = "00FF00"
GRAY_RGB = "808080"
WHITE_RGB = "FFFFFF"

CHECKIN_SHEET = 1
CHECKIN_COL_DATE = 0
CHECKIN_COL_ACCOUNT = 2
CHECKIN_COL_TYPE = 6
CHECKIN_COL_TIME = 8
CHECKIN_COL_STATUS = 9

HIDE_JOB_LEVELS = ["总监（业务）", "高级经理（业务）", "经理（业务）", "副经理（业务）"]

def parse_leave_data(file_path):
    df = pd.read_excel(file_path, header=0, skiprows=[1], dtype=str)
    df.columns = df.columns.str.strip()
    leaves = {}
    for _, row in df.iterrows():
        emp_id = str(row.iloc[4]).strip()
        date_range = row.iloc[8]
        if pd.isna(date_range):
            continue
        try:
            if "至" in date_range:
                start_str, end_str = date_range.split("至")
                start = pd.to_datetime(start_str.strip()).date()
                end = pd.to_datetime(end_str.strip()).date()
            else:
                start = end = pd.to_datetime(date_range.strip()).date()
        except:
            continue
        leave_type = row.iloc[7]
        total_hours = float(row.iloc[9])
        delta = (end - start).days + 1
        for i in range(delta):
            date = start + timedelta(days=i)
            leaves[(emp_id, date)] = (leave_type, total_hours, start, end)
    return leaves

def parse_checkin_data(file_path):
    df = pd.read_excel(file_path, sheet_name=CHECKIN_SHEET, dtype=str)
    checkins = {}
    for _, row in df.iterrows():
        date_str = row.iloc[CHECKIN_COL_DATE]
        try:
            date = pd.to_datetime(date_str.split()[0]).date()
        except:
            continue
        emp_id = str(row.iloc[CHECKIN_COL_ACCOUNT]).strip()
        punch_type = row.iloc[CHECKIN_COL_TYPE]
        punch_time = row.iloc[CHECKIN_COL_TIME]
        status = row.iloc[CHECKIN_COL_STATUS] if CHECKIN_COL_STATUS < len(row) else ""
        key = (emp_id, date)
        if key not in checkins:
            checkins[key] = {"上班": "", "下班": "", "外出": [], "异常": ""}
        if pd.notna(punch_time) and str(punch_time).strip():
            time_str = str(punch_time).strip()
            if time_str == "--":
                time_str = ""
            else:
                match = re.search(r'\d{1,2}:\d{2}', time_str)
                if match:
                    time_str = match.group()
        else:
            time_str = ""
        if punch_type == "上班":
            checkins[key]["上班"] = time_str
        elif punch_type == "下班":
            checkins[key]["下班"] = time_str
        elif "外出" in punch_type:
            if time_str:
                checkins[key]["外出"].append(time_str)
        if status and "缺卡" in status:
            checkins[key]["异常"] = status
    return checkins

def parse_remote_data(file_path):
    try:
        df = pd.read_excel(file_path, header=0, dtype=str)
        remote_dict = {}
        for _, row in df.iterrows():
            emp_id = str(row.iloc[1]).strip()
            date_range = row.iloc[6]
            leave_type = row.iloc[7]
            location = row.iloc[8]
            if pd.isna(date_range) or pd.isna(emp_id):
                continue
            try:
                if " - " in date_range:
                    start_str, end_str = date_range.split(" - ")
                    start = pd.to_datetime(start_str.strip()).date()
                    end = pd.to_datetime(end_str.strip()).date()
                else:
                    start = end = pd.to_datetime(date_range.strip()).date()
            except:
                continue
            delta = (end - start).days + 1
            for i in range(delta):
                date = start + timedelta(days=i)
                remote_dict[(emp_id, date)] = (leave_type, location)
        return remote_dict
    except Exception as e:
        print(f"读取远程办公文件失败: {e}")
        return {}

def get_date_from_cell_value(cell_value):
    if isinstance(cell_value, datetime):
        return cell_value.date()
    s = str(cell_value).strip()
    if '.' in s:
        s = s.split('.')[0]
    if len(s) == 8 and s.isdigit():
        return datetime.strptime(s, "%Y%m%d").date()
    elif '-' in s:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    return None

# ================== 新增：工时计算函数 ==================
def format_hours(cinfo):
    has_上班 = bool(cinfo.get("上班"))
    has_下班 = bool(cinfo.get("下班"))
    if has_上班 and has_下班:
        try:
            t1 = datetime.strptime(cinfo["上班"], "%H:%M")
            t2 = datetime.strptime(cinfo["下班"], "%H:%M")
            if t2 < t1:
                t2 += timedelta(days=1)
            diff_hours = (t2 - t1).total_seconds() / 3600
            if diff_hours > 0:
                return f" ({diff_hours:.1f}h)"
        except:
            pass
    return ""

def generate_cell_text(emp_id, date, leaves, checkins):
    if (emp_id, date) in leaves:
        leave_type, hours = leaves[(emp_id, date)]
        text = f"{leave_type}{hours}h"
        if (emp_id, date) in checkins:
            cinfo = checkins[(emp_id, date)]
            times = []
            if cinfo["上班"]:
                times.append(f"上班{cinfo['上班']}")
            if cinfo["下班"]:
                times.append(f"下班{cinfo['下班']}")
            if cinfo["外出"]:
                times.extend([f"外出{t}" for t in cinfo["外出"]])
            if times:
                text += " " + " ".join(times)
            text += format_hours(cinfo)  # 新增
        return text

    if (emp_id, date) in checkins:
        cinfo = checkins[(emp_id, date)]
        has_上班 = bool(cinfo["上班"])
        has_下班 = bool(cinfo["下班"])
        has_外出 = len(cinfo["外出"]) > 0

        if not has_上班 and not has_下班:
            text = "缺卡2次"
            if has_外出:
                text += " " + " ".join([f"外出{t}" for t in cinfo["外出"]])
            return text

        parts = []
        if has_上班:
            parts.append(f"上班{cinfo['上班']}")
        else:
            parts.append("缺卡1次")
        if has_下班:
            parts.append(f"下班{cinfo['下班']}")
        else:
            parts.append("缺卡1次")
        if has_外出:
            parts.extend([f"外出{t}" for t in cinfo["外出"]])
        text = " ".join(parts)
        text += format_hours(cinfo)  # 新增
        return text

    return "缺卡2次"

# ================== 颜色获取函数（与成功版本完全一致） ==================
def get_cell_color(cell):
    """获取单元格填充色（返回6位RGB十六进制，如 FF0000）。
    兼容 ARGB 格式（8位），自动剥离 alpha 通道。
    兼容 PatternFill 和 StyleProxy（BI模板常用）。
    """
    fill = cell.fill
    if not fill:
        return None

    def _extract_rgb(color_obj):
        """从 openpyxl Color 对象中提取 6 位 RGB"""
        if not color_obj:
            return None
        # 先获取 type，按类型分支处理（避免访问错误属性触发 descriptor 异常）
        ctype = getattr(color_obj, 'type', None)

        if ctype == 'rgb':
            rgb = getattr(color_obj, 'rgb', None)
            if rgb and isinstance(rgb, str) and rgb != '00000000':
                if len(rgb) == 8:
                    return rgb[2:].upper()
                return rgb.upper()

        elif ctype == 'indexed':
            try:
                idx = int(color_obj.indexed)
                if 0 <= idx < len(COLOR_INDEX):
                    color = COLOR_INDEX[idx]
                    if color:
                        if len(color) == 8:
                            return color[2:].upper()
                        return color.upper()
            except (ValueError, TypeError, IndexError):
                pass

        elif ctype == 'theme':
            theme_colors = {
                0: "000000", 1: "FFFFFF", 2: "FF0000",
                3: "00FF00", 4: "0000FF",
            }
            try:
                t = int(color_obj.theme)
                if t in theme_colors:
                    return theme_colors[t]
            except (ValueError, TypeError):
                pass

        return None

    # 尝试 fgColor（PatternFill / StyleProxy 都有）
    if hasattr(fill, 'fgColor'):
        result = _extract_rgb(fill.fgColor)
        if result:
            return result

    # 尝试 bgColor
    if hasattr(fill, 'bgColor'):
        result = _extract_rgb(fill.bgColor)
        if result:
            return result

    return None

def process_template_openpyxl(template_path, leaves, checkins, remote_dict, output_file, department="", start_date="", end_date=""):
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb["报表区"]

    date_row = 4
    first_date_col = 16
    date_cols = {}
    col = first_date_col
    empty_count = 0
    max_empty = 3
    while True:
        cell_value = ws.cell(row=date_row, column=col).value
        if cell_value is None or str(cell_value).strip() == "":
            empty_count += 1
            if empty_count >= max_empty:
                break
            col += 1
            continue
        empty_count = 0
        date = get_date_from_cell_value(cell_value)
        if date:
            date_cols[col] = date
        col += 1

    if not date_cols:
        print("错误：未能识别日期列，请检查模板！")
        sys.stdout.flush()
        return
    print(f"识别到的日期列数: {len(date_cols)}")
    sys.stdout.flush()

    last_row = 6
    for r in range(6, 501):
        val = ws.cell(row=r, column=2).value
        if val is not None and str(val).strip() != "":
            last_row = r
        else:
            empty_count = 0
            for i in range(r, min(r+5, 501)):
                v = ws.cell(row=i, column=2).value
                if v is None or str(v).strip() == "":
                    empty_count += 1
                else:
                    break
            if empty_count >= 5:
                break
    max_row = min(last_row + 2, 500)
    print(f"扫描行数: 6 到 {max_row}（动态识别）")
    sys.stdout.flush()

    # ===== 颜色规则（黑名单：跳过绿/灰/白/银/黑，其余填充） =====
    RED_HEX = "FF0000"
    YELLOW_HEX = "FFCC00"
    # 跳过色：绿=正常工作日 灰=非工作日 白=空白 黑=透明/无填充 浅灰/银=装饰色
    SKIP_COLORS = {"00FF00", "808080", "FFFFFF", "000000", "F0F0F0", "C0C0C0"}
    has_leave_data = len(leaves) > 0

    rows_to_hide = []
    modified_count = 0
    red_cells = []

    # ===== 第一遍扫描：收集每个员工的所有职级，处理合并单元格和多行情况 =====
    def get_merged_value(row, col):
        """读取单元格值，自动处理合并区域（返回左上角的值）"""
        cell = ws.cell(row=row, column=col)
        if cell.coordinate in ws.merged_cells:
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        return cell.value

    emp_job_levels = {}  # emp_id -> set of job levels
    emp_row_map = {}     # emp_id -> list of rows
    for row in range(6, max_row + 1):
        emp_id_raw = get_merged_value(row, 2)  # 姓名列
        if not emp_id_raw or str(emp_id_raw).strip() == "":
            continue
        emp_id = str(emp_id_raw).strip()
        emp_row_map.setdefault(emp_id, []).append(row)
        job_level = get_merged_value(row, 4)  # 职级列（处理合并单元格）
        if job_level:
            emp_job_levels.setdefault(emp_id, set()).add(str(job_level).strip())

    # 计算需要隐藏的员工：只要任意一个职级匹配 HIDE_JOB_LEVELS，就隐藏所有行
    employees_to_hide = set()
    for emp_id, levels in emp_job_levels.items():
        for level in levels:
            if any(hide in level for hide in HIDE_JOB_LEVELS):
                employees_to_hide.add(emp_id)
                break

    # 婚假/陪产假：必须一次性休完，包含周末和法定节假日
    LEAVE_FULL_DAYS_TYPES = {"婚假", "陪产假"}
    MAX_DAILY_HOURS = 8.0

    for row in range(6, max_row + 1):
        emp_cell = ws.cell(row=row, column=2)
        emp_id = None
        if emp_cell.coordinate in ws.merged_cells:
            for merged_range in ws.merged_cells.ranges:
                if emp_cell.coordinate in merged_range:
                    emp_id = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                    break
        else:
            emp_id = emp_cell.value
        if not emp_id or str(emp_id).strip() == "":
            continue
        emp_id = str(emp_id).strip()

        # 多职级员工：如果该员工任意一个职级命中隐藏规则，则隐藏本行
        if emp_id in employees_to_hide:
            rows_to_hide.append(row)
            continue

        # 休假天数分配
        emp_leave_records = []
        for (eid, date), (leave_type, total_hours, start, end) in leaves.items():
            if eid == emp_id:
                emp_leave_records.append((leave_type, total_hours, start, end))
        unique_records = []
        seen = set()
        for rec in emp_leave_records:
            key = (rec[1], rec[2], rec[3])
            if key not in seen:
                seen.add(key)
                unique_records.append(rec)

        leave_assignment = {}
        for leave_type, total_hours, start, end in unique_records:
            is_full_days_leave = leave_type in LEAVE_FULL_DAYS_TYPES

            if is_full_days_leave:
                # 婚假/陪产假：按日历天均摊（含周末和法定节假日），单日上限 8h
                total_days = 0
                for col2, date2 in date_cols.items():
                    if start <= date2 <= end:
                        total_days += 1
                if total_days > 0:
                    daily_hours = min(total_hours / total_days, MAX_DAILY_HOURS)
                    for col2, date2 in date_cols.items():
                        if start <= date2 <= end:
                            leave_assignment[(emp_id, date2)] = (leave_type, daily_hours)
            else:
                # 其他假期：按工作日均摊（跳过绿/灰/白/银/黑等非工作日）
                workday_count = 0
                for col2, date2 in date_cols.items():
                    if start <= date2 <= end:
                        color_hex = get_cell_color(ws.cell(row=row, column=col2))
                        if color_hex is None or color_hex not in SKIP_COLORS:
                            workday_count += 1
                if workday_count > 0:
                    daily_hours = total_hours / workday_count
                    for col2, date2 in date_cols.items():
                        if start <= date2 <= end:
                            color_hex = get_cell_color(ws.cell(row=row, column=col2))
                            if color_hex is None or color_hex not in SKIP_COLORS:
                                leave_assignment[(emp_id, date2)] = (leave_type, daily_hours)

        # 逐日填充
        for col, date in date_cols.items():
            cell = ws.cell(row=row, column=col)
            color_hex = get_cell_color(cell)

            # 黑名单判断：跳过不需要填充的颜色
            skip = False
            if color_hex is None:
                skip = True
            elif color_hex in SKIP_COLORS:
                skip = True
            # 无休假数据时跳过黄色（黄色代表休假，无休假文件 = 不需要填休假信息）
            elif not has_leave_data and color_hex == YELLOW_HEX:
                skip = True

            if skip:
                continue

            # 仅红色加入异常数据区（红色 = 缺卡异常）
            if color_hex == RED_HEX:
                red_cells.append((row, col))

            has_remote = remote_dict and (emp_id, date) in remote_dict
            remote_suffix = ""
            if has_remote:
                leave_type_remote, location = remote_dict[(emp_id, date)]
                if pd.isna(leave_type_remote) or str(leave_type_remote).strip() == "":
                    leave_type_remote = "远程工作"
                remote_suffix = f" {leave_type_remote}（{location}）"

            if (emp_id, date) in leave_assignment:
                leave_type, daily_hours = leave_assignment[(emp_id, date)]
                text = f"{leave_type}{daily_hours:.1f}h"
                if (emp_id, date) in checkins:
                    cinfo = checkins[(emp_id, date)]
                    times = []
                    if cinfo["上班"]:
                        times.append(f"上班{cinfo['上班']}")
                    if cinfo["下班"]:
                        times.append(f"下班{cinfo['下班']}")
                    if cinfo["外出"]:
                        times.extend([f"外出{t}" for t in cinfo["外出"]])
                    if times:
                        text += " " + " ".join(times)
                    text += format_hours(cinfo)
                if has_remote:
                    text += remote_suffix
                cell.value = text
                modified_count += 1
                continue

            text = generate_cell_text(emp_id, date, {}, checkins)
            if text is not None:
                if has_remote:
                    cell.value = text + remote_suffix
                else:
                    cell.value = text
                modified_count += 1

    for row in rows_to_hide:
        ws.row_dimensions[row].hidden = True

    for col in date_cols.keys():
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 12
        for row in range(6, max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True)

    # ===== 异常数据区 =====
    red_rows = set()
    red_cols = set()
    for (r, c) in red_cells:
        red_rows.add(r)
        red_cols.add(c)

    # 确定保留行：若员工任意一行有红色异常，该员工所有行都保留
    keep_rows = set()
    for emp_id, rows in emp_row_map.items():
        if any(r in red_rows for r in rows):
            for r in rows:
                keep_rows.add(r)

    if "异常数据区" in wb.sheetnames:
        del wb["异常数据区"]
    wb.create_sheet("异常数据区")
    ws_new = wb["异常数据区"]

    max_col = max(date_cols.keys()) if date_cols else 46

    # 复制表头（1-5行），所有列
    for r in range(1, 6):
        for c in range(1, max_col + 1):
            src = ws.cell(row=r, column=c)
            dst = ws_new.cell(row=r, column=c)
            dst.value = src.value
            if src.has_style:
                dst.font = src.font.copy()
                dst.border = src.border.copy()
                dst.fill = src.fill.copy()
                dst.number_format = src.number_format
                dst.protection = src.protection.copy()
                dst.alignment = src.alignment.copy()

    # 复制保留的数据行，只保留红色单元格内容，建立行号映射
    old_to_new_row = {}
    new_row = 6
    for old_row in sorted(keep_rows):
        for c in range(1, max_col + 1):
            src = ws.cell(row=old_row, column=c)
            dst = ws_new.cell(row=new_row, column=c)
            # 日期列中，非红色单元格清空
            if c in date_cols and (old_row, c) not in red_cells:
                dst.value = None
            else:
                dst.value = src.value
            if src.has_style:
                dst.font = src.font.copy()
                dst.border = src.border.copy()
                dst.fill = src.fill.copy()
                dst.number_format = src.number_format
                dst.protection = src.protection.copy()
                dst.alignment = src.alignment.copy()
        old_to_new_row[old_row] = new_row
        new_row += 1

    new_max_row = new_row - 1

    # 删除没有红色异常的列（保留基本信息列 1-15）
    for col in range(max_col, 15, -1):
        if col not in red_cols:
            ws_new.delete_cols(col)

    # 合并单元格
    # 1. 同一员工连续行的 B 列(2) 和 C 列(3) 合并
    for emp_id, rows in emp_row_map.items():
        kept_rows = sorted([r for r in rows if r in keep_rows])
        if len(kept_rows) <= 1:
            continue
        new_rows = [old_to_new_row[r] for r in kept_rows]
        ws_new.merge_cells(start_row=new_rows[0], start_column=2,
                           end_row=new_rows[-1], end_column=2)
        ws_new.merge_cells(start_row=new_rows[0], start_column=3,
                           end_row=new_rows[-1], end_column=3)

    # 2. 每一行合并 D+E 列(4+5)
    for r in range(6, new_max_row + 1):
        ws_new.merge_cells(start_row=r, start_column=4,
                           end_row=r, end_column=5)

    # 3. 每一行合并 F+G+H 列(6+7+8)
    for r in range(6, new_max_row + 1):
        ws_new.merge_cells(start_row=r, start_column=6,
                           end_row=r, end_column=8)

    # ===== 图表区 =====
    if "图表区" in wb.sheetnames:
        del wb["图表区"]
    wb.create_sheet("图表区", 0)  # 插入到最前面
    ws_chart = wb["图表区"]

    # 标题
    ws_chart["A1"] = f"{department} 考勤汇总统计"
    ws_chart["A1"].font = Font(bold=True, size=16)

    # 基本信息
    info = [
        ("统计期间", f"{start_date} 至 {end_date}"),
        ("部门名称", department),
        ("总人数", len(emp_row_map)),
        ("实际处理人数", len(emp_row_map) - len(employees_to_hide)),
        ("异常打卡人次", len(red_cells)),
        ("修改单元格数", modified_count),
    ]
    if remote_dict:
        info.append(("远程办公人次", len(remote_dict)))

    for i, (label, value) in enumerate(info, start=3):
        ws_chart.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws_chart.cell(row=i, column=2, value=value)

    # 请假类型统计
    leave_type_days = {}
    for (eid, date), (leave_type, total_hours, s, e) in leaves.items():
        leave_type_days[leave_type] = leave_type_days.get(leave_type, 0) + 1

    chart_row_start = 3 + len(info) + 1
    if leave_type_days:
        ws_chart.cell(row=chart_row_start, column=1, value="请假类型统计").font = Font(bold=True, size=12)
        chart_row_start += 1
        ws_chart.cell(row=chart_row_start, column=1, value="请假类型")
        ws_chart.cell(row=chart_row_start, column=2, value="人天")
        chart_row_start += 1
        row_idx = chart_row_start
        for leave_type, count in sorted(leave_type_days.items(), key=lambda x: -x[1]):
            ws_chart.cell(row=row_idx, column=1, value=leave_type)
            ws_chart.cell(row=row_idx, column=2, value=count)
            row_idx += 1

        # 柱状图
        chart = BarChart()
        chart.type = "col"
        chart.title = "请假类型分布"
        chart.y_axis.title = "人天"
        chart.x_axis.title = "请假类型"
        chart.style = 10
        data = Reference(ws_chart, min_col=2, min_row=chart_row_start - 1, max_row=row_idx - 1)
        cats = Reference(ws_chart, min_col=1, min_row=chart_row_start, max_row=row_idx - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 18
        chart.height = 10
        ws_chart.add_chart(chart, "D3")

    # 列宽调整
    ws_chart.column_dimensions["A"].width = 18
    ws_chart.column_dimensions["B"].width = 15

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_file))

    print(f"已生成考勤明细: {output_file}")
    print(f"隐藏职级行数: {len(rows_to_hide)}")
    print(f"实际修改单元格数: {modified_count}")
    print(f"异常数据区行数: {len([r for r in keep_rows if r >= 6])} 行, 列数: {len(red_cols)} 列")
    sys.stdout.flush()

def run_attendance_check(start_date, end_date, department, template_file, checkin_file, leave_file=None, remote_file=None):
    leaves = {}
    if leave_file and Path(leave_file).exists():
        print("正在读取休假数据...")
        sys.stdout.flush()
        leaves = parse_leave_data(leave_file)
        print(f"  共 {len(leaves)} 条请假记录（按天展开）")
        sys.stdout.flush()
    else:
        print("未提供休假文件，跳过")
        sys.stdout.flush()

    print("正在读取打卡数据...")
    sys.stdout.flush()
    checkins = parse_checkin_data(checkin_file)
    print(f"  共 {len(checkins)} 个员工-日期打卡记录")
    sys.stdout.flush()

    remote_dict = {}
    if remote_file and Path(remote_file).exists():
        print("正在读取远程办公数据...")
        sys.stdout.flush()
        remote_dict = parse_remote_data(remote_file)
        print(f"  共 {len(remote_dict)} 条远程办公记录")
        sys.stdout.flush()
    else:
        print("未提供远程办公文件，跳过")
        sys.stdout.flush()

    print("正在处理考勤模板...")
    sys.stdout.flush()

    base_dir = Path(__file__).parent
    output_dir = base_dir / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    start_str = f"{start.year}.{start.month}.{start.day}"
    end_str = f"{end.year}.{end.month}.{end.day}"
    filename = f"{department}考勤（{start_str}-{end_str}）.xlsx"
    output_file = output_dir / filename

    process_template_openpyxl(template_file, leaves, checkins, remote_dict, output_file, department, start_date, end_date)
    print("处理完成")
    sys.stdout.flush()

def main():
    parser = argparse.ArgumentParser(description="考勤核对工具")
    parser.add_argument("--start", required=True, help="起始日期，如 2026-04-01")
    parser.add_argument("--end", required=True, help="截止日期，如 2026-04-30")
    parser.add_argument("--dept", default="工程造价一部", help="部门名称")
    parser.add_argument("--template", required=True, help="考勤模板文件路径")
    parser.add_argument("--checkin", required=True, help="打卡日报文件路径")
    parser.add_argument("--leave", default=None, help="休假数据文件路径（可选）")
    parser.add_argument("--remote", default=None, help="远程办公文件路径（可选）")
    args = parser.parse_args()
    run_attendance_check(args.start, args.end, args.dept, args.template, args.checkin, args.leave, args.remote)

if __name__ == "__main__":
    main()
